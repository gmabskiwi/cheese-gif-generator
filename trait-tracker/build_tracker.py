"""NFT trait tracker — matrix layout: one row per trait, 6 variant slots across.

Rebuilds the tracker FROM SCRATCH (blank check state). For day-to-day
additions use add_traits.py instead, which appends to the existing file
without touching pasted art or checkmarks.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "NFT_Trait_Tracker.xlsx")

BG      = "0E1116"
CARD    = "1A1F27"
ROW_A   = "12161D"
ROW_B   = "161B23"
LINE    = "232A35"
TEXT    = "F5F7FA"
MUTED   = "8B94A3"
AMBER   = "FFB525"
GREEN   = "34D399"
RED     = "F87171"
GREEN_TINT = "16281F"   # image slot whose ✓ is set

AR = "Arial"
def f(size=10, bold=False, color=TEXT, italic=False):
    return Font(name=AR, size=size, bold=bold, color=color, italic=italic)
def fill(hexcode):
    return PatternFill("solid", fgColor=hexcode)

wb = Workbook()
ws = wb.active
ws.title = "Tracker"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = AMBER

for r in range(1, 101):
    for c in range(1, 20):
        ws.cell(row=r, column=c).fill = fill(BG)

# B cat, C trait, then 6 x (image col + check col) D..O, P done, Q notes
widths = {"A": 2, "B": 12, "C": 18,
          "D": 14, "E": 4, "F": 14, "G": 4, "H": 14, "I": 4,
          "J": 14, "K": 4, "L": 14, "M": 4, "N": 14, "O": 4,
          "P": 9, "Q": 24, "R": 2}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

IMG_COLS = ["D", "F", "H", "J", "L", "N"]
CHK_COLS = ["E", "G", "I", "K", "M", "O"]

# ---- hero ----
ws.row_dimensions[1].height = 14

ws.merge_cells("B2:Q2")
ws["B2"] = "NFT COLLECTION  ·  ART PROGRESS"
ws["B2"].font = f(9, bold=True, color=AMBER)
ws.row_dimensions[2].height = 14

ws.merge_cells("B3:Q3")
ws["B3"] = "TRAIT TRACKER"
ws["B3"].font = f(28, bold=True, color=TEXT)
ws["B3"].alignment = Alignment(vertical="center")
ws.row_dimensions[3].height = 38

ws.merge_cells("B4:Q4")
ws["B4"] = "One row per trait — all six versions across.  Male & female, in white, tanned and brown skin."
ws["B4"].font = f(10, color=MUTED)
ws.row_dimensions[4].height = 16

ws.row_dimensions[5].height = 10

# ---- stat cards (rows 6-7) ----
DATA_START, N_ROWS = 11, 60
DATA_END = DATA_START + N_ROWS - 1
GRID = f"$D${DATA_START}:$O${DATA_END}"
TRAITS = f"$C${DATA_START}:$C${DATA_END}"

ws.row_dimensions[6].height = 14
ws.row_dimensions[7].height = 30

def card(label_ref, value_ref, label, formula, color, numfmt=None):
    if ":" in label_ref:
        ws.merge_cells(label_ref)
        ws.merge_cells(value_ref)
    lc = ws[label_ref.split(":")[0]]
    vc = ws[value_ref.split(":")[0]]
    lc.value = label
    lc.font = f(8, bold=True, color=MUTED)
    lc.alignment = Alignment(horizontal="center", vertical="bottom")
    vc.value = formula
    vc.font = f(20, bold=True, color=color)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    if numfmt:
        vc.number_format = numfmt
    for ref in (label_ref, value_ref):
        a, b = (ref.split(":") + [ref.split(":")[0]])[:2]
        col_a, col_b = a[0], b[0]
        row = a[1:]
        for c in range(ord(col_a), ord(col_b) + 1):
            ws[f"{chr(c)}{row}"].fill = fill(CARD)

card("B6", "B7", "DONE", f'=SUMPRODUCT(({GRID}="✓")*1)', GREEN)
card("C6", "C7", "REMAINING", f'=6*COUNTA({TRAITS})-SUMPRODUCT(({GRID}="✓")*1)', RED)
card("D6:E6", "D7:E7", "TOTAL", f"=6*COUNTA({TRAITS})", TEXT)
card("F6:G6", "F7:G7", "% COMPLETE",
     f'=IF(COUNTA({TRAITS})=0,0,SUMPRODUCT(({GRID}="✓")*1)/(6*COUNTA({TRAITS})))',
     AMBER, "0%")
card("H6:O6", "H7:O7", "PROGRESS",
     '=REPT("█",ROUND($F$7*24,0))&REPT("░",24-ROUND($F$7*24,0))', AMBER)
ws["H7"].font = f(13, bold=True, color=AMBER)

# ---- instructions ----
ws.merge_cells("B8:Q8")
ws["B8"] = ("Each trait is one row.  Paste art onto its slot (normal paste, then drag it over the cell) "
            "and put a ✓ in the small box beside it when the image is final — the slot turns green.  "
            '"Gold Watch" and "Cigar" are examples — swap in your real traits.')
ws["B8"].font = f(9, italic=True, color=MUTED)
ws["B8"].alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[8].height = 28

# ---- two-tier header (rows 9-10) ----
accent_bottom = Border(bottom=Side(style="medium", color=AMBER))

ws.merge_cells("D9:I9")
ws["D9"] = "MALE"
ws.merge_cells("J9:O9")
ws["J9"] = "FEMALE"
for ref in ("D9", "J9"):
    ws[ref].font = f(10, bold=True, color=AMBER)
    ws[ref].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[9].height = 18

skins = ["WHITE", "TANNED", "BROWN"] * 2
for img_col, chk_col, skin in zip(IMG_COLS, CHK_COLS, skins):
    ws.merge_cells(f"{img_col}10:{chk_col}10")
    c = ws[f"{img_col}10"]
    c.value = skin
    c.font = f(9, bold=True, color=TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")

for label, ref in (("CATEGORY", "B"), ("TRAIT", "C"), ("DONE", "P"), ("NOTES", "Q")):
    ws.merge_cells(f"{ref}9:{ref}10")
    c = ws[f"{ref}9"]
    c.value = label
    c.font = f(9, bold=True, color=TEXT)
    c.alignment = Alignment(horizontal="center" if ref == "P" else "left", vertical="center")

for row in (9, 10):
    for col in "BCDEFGHIJKLMNOPQ":
        cell = ws[f"{col}{row}"]
        cell.fill = fill(CARD)
        if row == 10:
            cell.border = accent_bottom
ws.row_dimensions[10].height = 18

# ---- data rows ----
sep = Border(bottom=Side(style="thin", color=LINE))
chk_border = Border(bottom=Side(style="thin", color=LINE),
                    left=Side(style="thin", color=LINE))
for r in range(DATA_START, DATA_END + 1):
    ws.row_dimensions[r].height = 70
    band = fill(ROW_A) if r % 2 == 0 else fill(ROW_B)
    for col in "BCDEFGHIJKLMNOPQ":
        c = ws[f"{col}{r}"]
        c.fill = band
        c.border = chk_border if col in CHK_COLS else sep
        if col == "B":
            c.font = f(9, bold=True, color=MUTED)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        elif col == "C":
            c.font = f(11, bold=True, color=TEXT)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        elif col in CHK_COLS:
            c.font = f(12, bold=True, color=MUTED)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == "P":
            c.font = f(10, bold=True, color=MUTED)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == "Q":
            c.font = f(10, color=MUTED)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        else:  # image slots
            c.font = f(9, color=MUTED)
            c.alignment = Alignment(horizontal="center", vertical="center")
    # per-trait "n / 6" counter
    ws[f"P{r}"] = (f'=IF($C{r}="","",SUMPRODUCT(($D{r}:$O{r}="✓")*1)&" / 6")')

# example traits
ws[f"B{DATA_START}"] = "Left Arm"
ws[f"C{DATA_START}"] = "Gold Watch"
ws[f"E{DATA_START}"] = "✓"          # male-white done, as a demo
ws[f"B{DATA_START+1}"] = "Right Arm"
ws[f"C{DATA_START+1}"] = "Cigar"

# ✓ dropdown on the small boxes
dv = DataValidation(type="list", formula1='"✓"', allow_blank=True)
dv.error = "Tick with ✓ (or leave blank)"
dv.errorTitle = "Invalid value"
ws.add_data_validation(dv)
for col in CHK_COLS:
    dv.add(f"{col}{DATA_START}:{col}{DATA_END}")

# CF: ✓ boxes go green
chk_rng = " ".join(f"{col}{DATA_START}:{col}{DATA_END}" for col in CHK_COLS)
ws.conditional_formatting.add(
    chk_rng,
    CellIsRule(operator="equal", formula=['"✓"'],
               fill=fill(GREEN), font=Font(name=AR, size=12, bold=True, color=BG)))

# CF: image slot tints green when its ✓ is set
for img_col, chk_col in zip(IMG_COLS, CHK_COLS):
    ws.conditional_formatting.add(
        f"{img_col}{DATA_START}:{img_col}{DATA_END}",
        FormulaRule(formula=[f'${chk_col}{DATA_START}="✓"'], fill=fill(GREEN_TINT)))

# CF: DONE column celebrates a finished trait
ws.conditional_formatting.add(
    f"P{DATA_START}:P{DATA_END}",
    CellIsRule(operator="equal", formula=['"6 / 6"'],
               fill=fill(GREEN), font=Font(name=AR, size=10, bold=True, color=BG)))

ws.freeze_panes = f"D{DATA_START}"

wb.save(OUT)
print("saved", OUT)
