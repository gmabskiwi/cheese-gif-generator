"""Append traits to the tracker without touching existing art, checkmarks or formatting.

Usage:
    python add_traits.py "Left Arm: Rolex, Snake Tattoo" "Right Arm: Pinky Ring"

Each argument is one category followed by a comma-separated list of trait
names. Rows are written into the next empty pre-formatted rows of
NFT_Trait_Tracker.xlsx in this folder.
"""
import os
import sys
from openpyxl import load_workbook

TRACKER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "NFT_Trait_Tracker.xlsx")
DATA_START, DATA_END = 11, 70  # pre-formatted rows in the Tracker sheet


def parse_args(args):
    traits = []
    for arg in args:
        if ":" not in arg:
            sys.exit(f'Bad argument {arg!r} — expected "Category: name, name, ..."')
        category, names = arg.split(":", 1)
        for name in names.split(","):
            if name.strip():
                traits.append((category.strip(), name.strip()))
    return traits


def main():
    traits = parse_args(sys.argv[1:])
    if not traits:
        sys.exit("Nothing to add. " + __doc__)

    wb = load_workbook(TRACKER)
    ws = wb["Tracker"]

    existing = {str(ws[f"C{r}"].value).strip().lower()
                for r in range(DATA_START, DATA_END + 1) if ws[f"C{r}"].value}
    traits = [(c, t) for c, t in traits if t.lower() not in existing]

    row = DATA_START
    while row <= DATA_END and ws[f"C{row}"].value:
        row += 1
    free = DATA_END - row + 1
    if len(traits) > free:
        sys.exit(f"Only {free} formatted rows left but {len(traits)} traits to add — "
                 "extend the sheet with build_tracker.py styling first.")

    for category, trait in traits:
        ws[f"B{row}"] = category
        ws[f"C{row}"] = trait
        row += 1

    wb.save(TRACKER)
    for category, trait in traits:
        print(f"added  {category} · {trait}")
    print(f"{len(traits)} trait(s) added, {DATA_END - row + 1} formatted rows left.")


if __name__ == "__main__":
    main()
